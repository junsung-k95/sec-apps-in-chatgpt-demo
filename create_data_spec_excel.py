import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(name="맑은 고딕", bold=True, size=10)
cell_font = Font(name="맑은 고딕", size=10)
note_font = Font(name="맑은 고딕", size=10, italic=True, color="666666")
changed_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_subheader(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border

def style_cells(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

def auto_width(ws, max_col, min_width=12, max_width=50):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, min_width), max_width)

# ============================================================
# Sheet 1: 개요 (Overview)
# ============================================================
ws = wb.active
ws.title = "개요"

overview_headers = ["#", "데이터 영역", "설명", "현재 Mock 규모", "연결된 MCP Tool", "우선순위"]
ws.append(overview_headers)
style_header(ws, 1, len(overview_headers))

overview_data = [
    [1, "제품 카탈로그 (devices)", "기기 정보 + Trade-in 기준가 + 스펙 + 보조 테이블", "42개 기기 + 6개 보조 테이블", "start_tradein_appraisal, search_tradein_value, analyze_tradein_device", "필수"],
    [2, "Trade-in 가격 산정 테이블", "상태/지역/통신사별 가격 보정 계수 + 감가 항목 + Vision 판정 기준", "devices.json 내 6개 보조 테이블", "start_tradein_appraisal, search_tradein_value, analyze_tradein_device", "필수"],
    [3, "프로모션 캠페인 (promotions)", "할인, 번들, Trade-in 보너스, 구독 크레딧, 기업, 교육 등", "22개 캠페인", "get_promotions (미구현 → get_service_guidelines에서 참조)", "필수"],
    [4, "Galaxy Club 구독 플랜 (plans)", "구독 플랜 비교, 비용 시뮬레이션, 지역별 가격, FAQ, 가입 절차", "3개 플랜 + 7건 비용비교 + 10건 FAQ", "get_galaxy_club_info, compare_galaxy_club_cost", "필수"],
    [5, "Samsung Care+ 보험 (care_plus)", "보험 플랜, Late Enrollment, Vision 검사, 클레임 절차, 서비스센터", "2개 플랜 + 7개 기기카테고리 가격 + 10건 FAQ", "get_care_plus_info, check_care_plus_eligibility", "필수"],
]
for row in overview_data:
    ws.append(row)
style_cells(ws, 2, len(overview_data) + 1, len(overview_headers))

ws.append([])
r = ws.max_row + 1
ws.cell(row=r, column=1, value="현재 등록된 MCP Tool 목록 (9개)")
ws.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["#", "Tool Name", "설명", "Read-only", "위젯 바인딩", ""], 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 6)

tools = [
    [1, "get_service_guidelines", "서비스 가이드라인 (대화 시작 시 호출)", "Y", "없음"],
    [2, "get_galaxy_club_info", "Galaxy Club 구독 플랜 정보", "Y", "galaxyClubWidget"],
    [3, "start_tradein_appraisal", "Trade-in 견적 시작 (appraisalStore에 기록)", "N", "tradeinWidget"],
    [4, "get_tradein_result", "Trade-in 견적 결과 조회", "Y", "tradeinWidget"],
    [5, "search_tradein_value", "Trade-in 가격 범위 조회 (위젯에서 tools/call 호출 가능)", "Y", "tradeinWidget (visibility: model+app)"],
    [6, "get_care_plus_info", "Care+ 보험 플랜 정보 + Late Enrollment 판단", "Y", "carePlusWidget"],
    [7, "check_care_plus_eligibility", "Care+ Late Enrollment Vision 기반 적격성 판단", "N", "carePlusWidget"],
    [8, "compare_galaxy_club_cost", "Galaxy Club vs 일시불 비용 비교 (위젯에서 tools/call 호출 가능)", "Y", "comparisonWidget (visibility: model+app)"],
    [9, "analyze_tradein_device", "Vision 기반 Trade-in 재산정 (appraisal 업데이트)", "N", "tradeinWidget"],
]
start_r = r + 1
for row in tools:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_cells(ws, start_r, r, 6)

ws.append([])
ws.append(["", "※ 가격 산출 공식: (base_value × region_multiplier) + condition_adj - functional_deduction - cosmetic_deduction + carrier_adj + promo_bonus"])
ws.cell(row=ws.max_row, column=2).font = note_font
ws.append(["", "※ 서버에 요청/응답 로깅 구현됨 (registerAppToolWithLogging 래퍼)"])
ws.cell(row=ws.max_row, column=2).font = note_font
ws.append(["", "※ 6개 지역 통화 지원: US($), KR(원), UK(£), DE(€), JP(¥), SG(S$)"])
ws.cell(row=ws.max_row, column=2).font = note_font

auto_width(ws, len(overview_headers))

# ============================================================
# Sheet 2: 제품 카탈로그 (devices)
# ============================================================
ws2 = wb.create_sheet("제품 카탈로그")

headers2 = ["필드명", "타입", "필수", "설명", "예시 값", "비고"]
ws2.append(headers2)
style_header(ws2, 1, len(headers2))

devices_data = [
    ["device_id", "string (PK)", "Y", "기기 고유 식별자", "DEV-SM-S926B", "삼성 내부 모델 코드 기준"],
    ["model", "string", "Y", "모델명 (검색 키)", "Galaxy S25 Ultra", "사용자 검색에 사용"],
    ["series", "string", "Y", "시리즈 분류", "Galaxy S, Galaxy Z", "필터링 용도"],
    ["category", "enum", "Y", "카테고리", "smartphone / tablet / watch / earbuds / notebook / wearable", "위젯 필터링 용도"],
    ["generation", "integer", "Y", "세대 번호", "25", ""],
    ["storage_options", "string[]", "Y", "용량 옵션", '["256GB", "512GB", "1TB"]', ""],
    ["base_values", "map<string, number>", "Y", "용량별 Trade-in 기준가 (USD)", '{"256GB": 680, "512GB": 750}', "★ Trade-in 핵심 데이터"],
    ["msrp", "map<string, number>", "Y", "용량별 정가 (USD)", '{"256GB": 1299, "512GB": 1419}', "Galaxy Club 비용 비교에 사용"],
    ["colors", "string[]", "N", "색상 옵션", '["Titanium Black", "Titanium Gray"]', ""],
    ["release_date", "date", "Y", "출시일", "2025-01-22", "Care+ Late Enrollment 판단 기준"],
    ["release_year", "integer", "Y", "출시 연도", "2025", "위젯 표시용"],
    ["end_of_support_date", "date", "N", "지원 종료일", "2032-01-22", ""],
    ["status", "enum", "Y", "판매 상태", "active / discontinued", ""],
    ["trade_in_eligible", "boolean", "Y", "Trade-in 대상 여부", "true / false", "false이면 Trade-in 불가"],
    ["specs.display", "string", "N", "디스플레이 스펙", '6.9" Dynamic AMOLED 2X, QHD+, 120Hz', "위젯 상세 표시용"],
    ["specs.processor", "string", "N", "프로세서", "Snapdragon 8 Elite for Galaxy", ""],
    ["specs.ram", "string", "N", "RAM", "12GB", ""],
    ["specs.battery", "string", "N", "배터리", "5000mAh", ""],
    ["specs.camera", "string", "N", "카메라 구성", "200MP + 50MP + 10MP + 50MP", ""],
    ["specs.os", "string", "N", "운영체제", "One UI 7 / Android 15", ""],
    ["specs.water_resistance", "string", "N", "방수 등급", "IP68", ""],
    ["specs.weight_g", "number", "N", "무게 (g)", "218", ""],
    ["specs.dimensions_mm", "string", "N", "크기 (mm)", "162.8 x 77.6 x 8.2", ""],
]
for row in devices_data:
    ws2.append(row)
style_cells(ws2, 2, len(devices_data) + 1, len(headers2))

r = ws2.max_row + 2
ws2.cell(row=r, column=1, value="현재 Mock 기기 라인업 (42개)")
ws2.cell(row=r, column=1).font = subheader_font

lineup = [
    ["Galaxy S Series (12)", "S25 Ultra / S25+ / S25 / S24 Ultra / S24+ / S24 / S23 Ultra / S23+ / S23 / S22 Ultra / S22+ / S22"],
    ["Galaxy Z Series (6)", "Z Fold6 / Z Flip6 / Z Fold5 / Z Flip5 / Z Fold4 / Z Flip4"],
    ["Galaxy Tab S (6)", "Tab S10 Ultra / S10+ / S10 / Tab S9 Ultra / S9+ / S9"],
    ["Galaxy Watch (5)", "Watch Ultra / Watch7 / Watch7 Classic / Watch6 Classic / Watch6"],
    ["Galaxy Buds (4)", "Buds3 Pro / Buds3 / Buds2 Pro / Buds FE"],
    ["Galaxy Book (6)", "Book4 Ultra / Book4 Pro / Book4 Pro 360 / Book4 360 / Book4 / Book Go"],
    ["Galaxy A Series (3)", "A55 / A35 / A25"],
    ["Galaxy Ring (1)", "Ring"],
]
for item in lineup:
    r += 1
    ws2.cell(row=r, column=1, value=item[0]).font = cell_font
    ws2.cell(row=r, column=2, value=item[1]).font = cell_font

auto_width(ws2, len(headers2))

# ============================================================
# Sheet 3: Trade-in 가격 산정
# ============================================================
ws3 = wb.create_sheet("Trade-in 가격 산정")

# Section: condition_multipliers
ws3.append(["[1] 상태별 계수 (condition_multipliers)"])
style_subheader(ws3, 1, 4)
ws3.append(["상태", "계수", "설명", "비고"])
style_header(ws3, 2, 4)
cond_data = [
    ["excellent", 1.00, "새것과 동일", ""],
    ["good", 0.85, "미세한 스크래치", ""],
    ["fair", 0.65, "눈에 보이는 사용감", ""],
    ["poor", 0.40, "심한 파손", ""],
]
for row in cond_data:
    ws3.append(row)
style_cells(ws3, 3, 6, 4)

ws3.append([])

# Section: region_multipliers
r = ws3.max_row + 1
ws3.cell(row=r, column=1, value="[2] 지역별 계수 (region_multipliers)")
style_subheader(ws3, r, 6)
r += 1
for i, h in enumerate(["지역", "계수", "통화", "기호", "환율 (대 USD)", "세금 포함"], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 6)

region_data = [
    ["US", 1.00, "USD", "$", "-", "N"],
    ["KR", 0.92, "KRW", "₩", 1350, "Y"],
    ["UK", 0.95, "GBP", "£", 0.79, "Y"],
    ["DE", 0.93, "EUR", "€", 0.92, "Y"],
    ["JP", 0.88, "JPY", "¥", 155, "Y"],
    ["SG", 0.96, "SGD", "S$", 1.35, "Y"],
]
start_r = r + 1
for row in region_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws3.cell(row=r, column=i, value=v)
style_cells(ws3, start_r, r, 6)

ws3.append([])

# Section: carrier_adjustments
r = ws3.max_row + 1
ws3.cell(row=r, column=1, value="[3] 통신사별 가격 조정 (carrier_adjustments)")
style_subheader(ws3, r, 4)
r += 1
for i, h in enumerate(["통신사", "조정액 (USD)", "비고", ""], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 4)

carrier_data = [
    ["unlocked", 0, "자급제"],
    ["samsung_direct", "+10", "삼성닷컴 직구매"],
    ["att / verizon", "-15", "미국 통신사"],
    ["tmobile", "-10", "미국 통신사"],
    ["skt / kt / lgu", "-10", "한국 통신사"],
    ["docomo", "-20", "일본 통신사"],
    ["other_locked", "-25", "기타 잠금"],
]
start_r = r + 1
for row in carrier_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws3.cell(row=r, column=i, value=v)
style_cells(ws3, start_r, r, 4)

ws3.append([])

# Section: functional_issues
r = ws3.max_row + 1
ws3.cell(row=r, column=1, value="[4] 기능 문제 감가 (functional_issues_deductions)")
style_subheader(ws3, r, 4)
r += 1
for i, h in enumerate(["항목", "감가액 (USD)", "설명", ""], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 4)

func_data = [
    ["screen_burn", -50, "화면 번인"],
    ["camera_issues", -40, "카메라 불량"],
    ["charging_issue", -35, "충전 문제"],
    ["fingerprint_sensor", -30, "지문인식 불량"],
    ["battery_issue", -30, "배터리 문제"],
    ["speaker_issue", -25, "스피커 불량"],
    ["microphone", -25, "마이크 불량"],
    ["button_issue", -20, "버튼 불량"],
    ["face_recognition", -20, "안면인식 불량"],
    ["connectivity", -15, "연결성 문제 (WiFi/BT)"],
    ["gps", -15, "GPS 불량"],
]
start_r = r + 1
for row in func_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws3.cell(row=r, column=i, value=v)
style_cells(ws3, start_r, r, 4)

ws3.append([])

# Section: cosmetic_issues
r = ws3.max_row + 1
ws3.cell(row=r, column=1, value="[5] 외관 문제 감가 (cosmetic_issues_deductions)")
style_subheader(ws3, r, 4)
r += 1
for i, h in enumerate(["항목", "감가액 (USD)", "설명", ""], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 4)

cosm_data = [
    ["screen_cracks", -100, "화면 깨짐"],
    ["back_cracks", -60, "후면 깨짐"],
    ["hinge_wear", -40, "힌지 마모 (폴더블)"],
    ["dents", -25, "찍힘"],
    ["s_pen_damage", -20, "S Pen 파손"],
    ["port_damage", -20, "충전 포트 파손"],
    ["screen_scratches", -15, "화면 스크래치"],
    ["discoloration", -15, "변색"],
    ["bezel_damage", -15, "베젤 손상"],
    ["back_scratches", -10, "후면 스크래치"],
]
start_r = r + 1
for row in cosm_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws3.cell(row=r, column=i, value=v)
style_cells(ws3, start_r, r, 4)

ws3.append([])

# Section: vision criteria
r = ws3.max_row + 1
ws3.cell(row=r, column=1, value="[6] AI Vision 판정 기준 (vision_assessment_criteria)")
style_subheader(ws3, r, 5)
r += 1
for i, h in enumerate(["부위", "excellent", "good", "fair", "poor"], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 5)

vision_data = [
    ["화면 (screen_condition)", "no_scratches", "light_scratches", "visible_scratches", "cracked"],
    ["바디 (body_condition)", "pristine", "minor_wear", "dents_scratches", "major_damage"],
    ["카메라 (camera_condition)", "clear", "minor_smudge", "scratched", "cracked"],
    ["힌지 (hinge_condition, 폴더블)", "tight_no_play", "slight_wobble", "loose_creaky", "damaged_misaligned"],
]
start_r = r + 1
for row in vision_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws3.cell(row=r, column=i, value=v)
style_cells(ws3, start_r, r, 5)

auto_width(ws3, 6)

# ============================================================
# Sheet 4: 프로모션 캠페인
# ============================================================
ws4 = wb.create_sheet("프로모션 캠페인")

headers4 = ["필드명", "타입", "필수", "설명", "예시 값", "비고"]
ws4.append(headers4)
style_header(ws4, 1, len(headers4))

promo_data = [
    ["id", "string (PK)", "Y", "프로모션 ID", "promo-001", ""],
    ["campaign_id", "string", "Y", "캠페인 코드", "CAMP-2026-Q1-S25", "내부 관리용"],
    ["title", "string", "Y", "프로모션 제목", "Galaxy S25 Ultra Launch Offer", "위젯 카드 타이틀"],
    ["description", "string", "Y", "상세 설명", "Trade-in 시 $200 즉시 크레딧", "위젯 카드 본문"],
    ["discount_amount", "number", "Y", "할인 금액 또는 비율", "200", "discount_type에 따라 해석 달라짐"],
    ["discount_type", "enum", "Y", "할인 유형 (아래 별도 표 참조)", "fixed", "fixed/percentage/bundle 등 11종"],
    ["eligible_devices", "string[]", "Y", "적용 대상 기기 목록", '["Galaxy S25", "Galaxy S25+"]', "devices.model과 매칭"],
    ["device_category", "enum", "Y", "카테고리 필터", "smartphone / tablet / notebook / watch / earbuds / wearable / all / cross", ""],
    ["target_audience", "enum", "Y", "타겟 고객군", "all / student / enterprise / creator / upgrade / new_subscriber / budget / ecosystem_member / late_enrollee / members_gold_platinum", ""],
    ["status", "enum", "Y", "상태", "active / expired / scheduled", ""],
    ["valid_from", "date", "Y", "시작일", "2025-01-22", ""],
    ["valid_until", "date", "Y", "종료일", "2026-04-30", ""],
    ["conditions", "string[]", "N", "적용 조건", '["Trade-in required", "Online only"]', "위젯에 조건 표시"],
    ["stackable", "boolean", "Y", "다른 프로모션과 중복 적용 가능 여부", "true / false", ""],
    ["max_redemptions", "integer|null", "N", "최대 사용 수 (null=무제한)", "5000", ""],
    ["current_redemptions", "integer", "N", "현재 사용 수", "3241", "재고/인기도 참고"],
    ["cta_text", "string", "Y", "CTA 버튼 텍스트", "Shop Now", ""],
    ["cta_url", "string", "Y", "CTA 버튼 링크", "https://samsung.com/promo/...", ""],
    ["image_url", "string", "N", "프로모션 이미지 URL", "https://images.samsung.com/...", "위젯 카드 이미지"],
    ["priority", "integer", "Y", "표시 우선순위 (1=최상위)", "1", "위젯 정렬 기준"],
    ["features", "string[]", "N", "추가 기능 목록 (AI/Enterprise 프로모션)", '["Knox Vault", "Zero-touch enrollment"]', "promo-009, 013에서 사용"],
    ["modules", "string[]", "N", "교육 모듈 목록 (Academy 프로모션)", '["AI Photo Editing Basics", ...]', "promo-010에서 사용"],
    ["bundle_options", "string[]", "N", "번들 옵션 목록 (Campus 프로모션)", '["Study Bundle: ...", "Research Bundle: ..."]', "promo-011에서 사용"],
    ["created_at", "datetime", "N", "생성 일시", "2025-01-10T00:00:00Z", ""],
    ["updated_at", "datetime", "N", "수정 일시", "2026-03-01T00:00:00Z", ""],
]
for row in promo_data:
    ws4.append(row)
style_cells(ws4, 2, len(promo_data) + 1, len(headers4))

ws4.append([])
r = ws4.max_row + 1
ws4.cell(row=r, column=1, value="discount_type 유형 정의 (11종)")
ws4.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["유형", "의미", "discount_amount 해석", "사용 프로모션", "", ""], 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, 6)

dtype_data = [
    ["fixed", "고정 금액 할인", "$200 할인", "promo-001,005,007,008,012,014,019,020"],
    ["percentage", "비율 할인", "15% 할인", "promo-003,004,016,017"],
    ["bundle", "번들 증정", "$249 상당 제품 무료 증정", "promo-002,018"],
    ["trade_in_max", "Trade-in 최대 보상", "최대 $1200 크레딧", "promo-006"],
    ["percentage_max", "최대 비율 할인", "최대 30% 할인", "promo-011"],
    ["percentage_bonus", "Trade-in 추가 보너스", "기존 Trade-in에 +20%", "promo-015"],
    ["subscription_credit", "구독 크레딧", "$30 상당 구독 무료", "promo-009"],
    ["subscription_bonus", "구독 가입 보너스", "$150 상당 혜택", "promo-021"],
    ["fee_waiver", "수수료 면제", "수수료 $0", "promo-022"],
    ["free_program", "무료 프로그램", "금액 없음", "promo-010"],
    ["enterprise", "기업 맞춤", "별도 협의", "promo-013"],
]
start_r = r + 1
for row in dtype_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws4.cell(row=r, column=i, value=v)
style_cells(ws4, start_r, r, 6)

# target_audience 유형
ws4.append([])
r = ws4.max_row + 1
ws4.cell(row=r, column=1, value="target_audience 유형")
ws4.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["유형", "설명", "사용 프로모션", "", "", ""], 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, 6)

audience_data = [
    ["all", "전체 고객", "promo-001~003,005,009,010,014,016,019,020"],
    ["student", "학생 (.edu 인증)", "promo-004,007,011"],
    ["enterprise", "기업 고객 (10대 이상)", "promo-013"],
    ["creator", "크리에이터/전문가", "promo-008"],
    ["upgrade", "기존 기기 업그레이드 고객", "promo-006,015"],
    ["new_subscriber", "Galaxy Club 신규 가입자", "promo-021"],
    ["budget", "가성비 고객", "promo-012"],
    ["ecosystem_member", "Galaxy 기기 3대 이상 보유자", "promo-005"],
    ["late_enrollee", "Care+ 60일 초과 가입자", "promo-022"],
    ["members_gold_platinum", "Galaxy Members 골드/플래티넘", "promo-017"],
]
start_r = r + 1
for row in audience_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws4.cell(row=r, column=i, value=v)
style_cells(ws4, start_r, r, 6)

auto_width(ws4, len(headers4))

# ============================================================
# Sheet 5: Galaxy Club 구독 플랜
# ============================================================
ws5 = wb.create_sheet("Galaxy Club 플랜")

headers5 = ["필드명", "타입", "필수", "설명", "예시 값", "비고"]
ws5.append(headers5)
style_header(ws5, 1, len(headers5))

plan_data = [
    ["id", "string (PK)", "Y", "플랜 ID", "ngc-basic / ngc-premium / ngc-family", ""],
    ["name", "string", "Y", "플랜명", "Basic Plan / Premium Plan / Family Plan", "위젯 카드 타이틀"],
    ["tier", "integer", "Y", "플랜 등급", "1 / 2 / 3", "정렬 기준"],
    ["monthly_price", "number", "Y", "월정액 (USD)", "35 / 55 / 89", ""],
    ["annual_price", "number", "Y", "연간 합계 (USD)", "378 / 594 / 961", "비용 비교용"],
    ["currency", "string", "Y", "기준 통화", "USD", ""],
    ["upgrade_cycle_months", "integer", "Y", "업그레이드 주기 (개월)", "24 / 12 / 12", "핵심 차별점"],
    ["included_devices", "string[]", "Y", "포함 기기 라인", '["Galaxy S Series (S25, S25+)", ...]', "플랜별 상이"],
    ["excluded_devices", "string[]", "Y", "미포함 기기", '["Galaxy S Ultra models", ...]', "Basic은 Ultra/Z 미포함"],
    ["benefits", "string[]", "Y", "혜택 목록 (한글)", '["24개월마다 기기 업그레이드", ...]', "위젯에서 최대 6개 표시"],
    ["not_included", "string[]", "N", "미포함 혜택", '["Ultra/Premium devices", ...]', "위젯에서 미사용"],
    ["terms.minimum_commitment_months", "integer", "Y", "최소 약정 기간 (개월)", "12", ""],
    ["terms.early_termination_fee", "number", "Y", "중도 해지 수수료 (USD)", "150 / 200 / 300", ""],
    ["terms.device_return_required", "boolean", "Y", "기기 반납 필수 여부", "true", ""],
    ["terms.damage_fee_max", "number", "Y", "파손 시 최대 수수료 (USD)", "250 / 350 / 350", ""],
    ["terms.late_payment_fee", "number", "Y", "연체 수수료 (USD)", "15 / 20 / 25", ""],
    ["terms.device_buyout_eligible_after_months", "integer", "Y", "바이아웃 가능 시점 (개월)", "12", ""],
    ["terms.buyout_residual_percentage", "number", "Y", "바이아웃 잔존가율 (%)", "40 / 35 / 35", ""],
    ["terms.max_lines", "integer", "N", "최대 라인 수 (Family만)", "4", "Family 플랜 전용"],
    ["terms.additional_line_fee", "number", "N", "추가 라인 수수료 (Family만)", "15", "Family 플랜 전용"],
    ["subscriber_count", "integer", "N", "현재 가입자 수", "48230 / 31450 / 12780", ""],
    ["status", "enum", "Y", "상태", "active", ""],
]
for row in plan_data:
    ws5.append(row)
style_cells(ws5, 2, len(plan_data) + 1, len(headers5))

# Regional pricing
ws5.append([])
r = ws5.max_row + 1
ws5.cell(row=r, column=1, value="지역별 가격 (regional_pricing)")
ws5.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["지역", "Basic/월", "Premium/월", "Family/월", "통화", ""], 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 6)

rp_data = [
    ["US", "$35", "$55", "$89", "USD"],
    ["KR", "₩39,900", "₩62,900", "₩99,900", "KRW"],
    ["UK", "£29", "£45", "£72", "GBP"],
    ["DE", "€33", "€52", "€84", "EUR"],
    ["JP", "¥4,980", "¥7,980", "¥12,980", "JPY"],
    ["SG", "S$48", "S$75", "S$120", "SGD"],
]
start_r = r + 1
for row in rp_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws5.cell(row=r, column=i, value=v)
style_cells(ws5, start_r, r, 6)

# Cost comparison
ws5.append([])
r = ws5.max_row + 1
ws5.cell(row=r, column=1, value="비용 비교 데이터 (cost_comparison) — 7건")
ws5.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["기기", "정가", "플랜", "12개월 총액", "절감액", "비고"], 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 6)

cost_data = [
    ["S25 Ultra 256GB", "$1,299", "Premium ($55/월)", "$660", "$639", "Care+ $215 가치 포함"],
    ["S25+ 256GB", "$999", "Basic ($35/월)", "$420", "$579", "Care+ $144 가치 포함"],
    ["Z Fold6 256GB", "$1,899", "Premium ($55/월)", "$660", "$1,239", "Care+ $215 가치 포함"],
    ["Z Flip6 256GB", "$1,099", "Premium ($55/월)", "$660", "$439", "Care+ $215 가치 포함"],
    ["S25 Ultra + S23U Trade-in", "$1,299", "Premium ($55/월)", "$660", "$189", "일시불 $849 (Trade-in $450 적용 후)"],
    ["Tab S10 Ultra 256GB", "$1,199", "Premium ($55/월)", "$660", "$539", "Care+ $215 가치 포함"],
    ["S25 128GB", "$799", "Basic ($35/월)", "$840 (24개월)", "-$41", "순수 기기비 비슷. Care+ 포함 시 $247 절약"],
]
start_r = r + 1
for row in cost_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws5.cell(row=r, column=i, value=v)
style_cells(ws5, start_r, r, 6)

# Lifecycle
ws5.append([])
r = ws5.max_row + 1
ws5.cell(row=r, column=1, value="라이프사이클 (lifecycle.stages) — 6단계")
ws5.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["단계", "시점", "라벨", "설명", "", ""], 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 6)

lc_data = [
    [1, "M0", "가입 신청", "플랜 선택, 크레딧 체크, 결제 설정"],
    [2, "M1", "기기 수령", "새 Galaxy 기기 수령, Care+ 자동 적용"],
    [3, "M1-11", "기기 사용", "Care+ 보장, 멤버 혜택 이용, Galaxy AI 활용"],
    [4, "M6", "중간 점검", "얼리 업그레이드 옵션 안내 (Premium/Family)"],
    [5, "M12", "업그레이드 선택", "반납 후 최신 기기로 업그레이드, 바이아웃, 또는 연장"],
    [6, "M13+", "새 사이클", "최신 기기로 새로운 12개월 시작"],
]
start_r = r + 1
for row in lc_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws5.cell(row=r, column=i, value=v)
style_cells(ws5, start_r, r, 6)

# FAQ
ws5.append([])
r = ws5.max_row + 1
ws5.cell(row=r, column=1, value="FAQ (10건, 전체 한글)")
ws5.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["ID", "카테고리", "질문", "답변 (요약)", "", ""], 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 6)

faq_data = [
    ["faq-001", "general", "현재 전화번호를 유지할 수 있나요?", "기기 구독 서비스, 통신사 무관"],
    ["faq-002", "upgrade", "업그레이드 시 기존 기기는?", "무료 배송 라벨로 반납"],
    ["faq-003", "billing", "언제든 해지 가능?", "12개월 후 가능, 조기 해지 위약금"],
    ["faq-004", "protection", "기기 파손 시?", "Care+ 포함, 심한 파손 시 수수료"],
    ["faq-005", "ownership", "바이아웃 가능?", "12개월 후 잔존가 기준 매입"],
    ["faq-006", "accessories", "액세서리 포함?", "Premium/Family: 번들 $100"],
    ["faq-007", "upgrade", "조기 업그레이드?", "비례 수수료, Premium 6개월 후 할인"],
    ["faq-008", "general", "일반 구매와 차이?", "월정액 vs 소유, Care+/업그레이드 포함"],
    ["faq-009", "trade_in", "Trade-in 적용?", "반납형이라 별도 Trade-in과 다름"],
    ["faq-010", "billing", "결제 방법?", "신용카드/체크카드/Samsung Pay"],
]
start_r = r + 1
for row in faq_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws5.cell(row=r, column=i, value=v)
style_cells(ws5, start_r, r, 6)

# Enrollment steps
ws5.append([])
r = ws5.max_row + 1
ws5.cell(row=r, column=1, value="가입 절차 (enrollment_steps) — 5단계, 전체 한글")
ws5.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["단계", "타이틀", "설명", "소요 시간 (분)", "", ""], 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 6)

es_data = [
    [1, "플랜 선택", "플랜을 선택하세요 (Basic, Premium, Family)", "3분"],
    [2, "기기 선택", "대상 기기 라인업에서 첫 기기를 선택하세요", "5분"],
    [3, "신용 조회", "간편 신용 조회 (소프트 풀, 신용 점수에 영향 없음)", "2분"],
    [4, "주문 확인", "주문 확인 및 월 결제 설정", "3분"],
    [5, "기기 수령", "2-3 영업일 내 기기 수령, 풀 보장 적용!", "-"],
]
start_r = r + 1
for row in es_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws5.cell(row=r, column=i, value=v)
style_cells(ws5, start_r, r, 6)

auto_width(ws5, len(headers5))

# ============================================================
# Sheet 6: Samsung Care+
# ============================================================
ws6 = wb.create_sheet("Samsung Care+")

headers6 = ["필드명", "타입", "필수", "설명", "예시 값", "비고"]
ws6.append(headers6)
style_header(ws6, 1, len(headers6))

care_data = [
    ["id", "string (PK)", "Y", "플랜 ID", "care-basic / care-premium", ""],
    ["name", "string", "Y", "플랜명", "Care+ Basic / Care+ Premium", ""],
    ["tier", "integer", "Y", "플랜 등급", "1 / 2", ""],
    ["monthly_price", "number", "Y", "월 보험료 (USD, 플래그십 기준)", "11.99 / 17.99", "기기 카테고리별 차등"],
    ["annual_price", "number", "Y", "연간 보험료 (USD)", "129 / 199", ""],
    ["coverage", "string[]", "Y", "보장 항목 (한글)", '["우발적 파손 수리 (연 2회)", ...]', "위젯 체크리스트"],
    ["coverage_en", "string[]", "N", "보장 항목 (영문)", '["Accidental damage repair", ...]', "다국어 지원용"],
    ["deductible.screen_repair", "number", "Y", "화면 수리 자기부담금", "$29 / $0", "Premium $0이 핵심 셀링포인트"],
    ["deductible.back_glass_repair", "number", "Y", "후면 유리 수리 자기부담금", "$29 / $0", ""],
    ["deductible.other_repair", "number", "Y", "기타 수리 자기부담금", "$99 / $49", ""],
    ["deductible.replacement", "number", "N", "교체 자기부담금 (Premium만)", "- / $149", ""],
    ["deductible.lost_stolen_replacement", "number", "N", "분실 교체 자기부담금 (Premium만)", "- / $249", ""],
    ["deductible.battery_replacement", "number", "Y", "배터리 교체 자기부담금", "$0 / $0", ""],
    ["deductible.hinge_repair", "number", "N", "힌지 수리 자기부담금 (Premium만)", "- / $0", "폴더블 전용"],
    ["coverage_limit_per_claim", "number", "Y", "클레임당 보장 한도", "$1,000 / $2,000", ""],
    ["coverage_limit_annual", "number", "Y", "연간 보장 한도", "$2,500 / $5,000", ""],
    ["eligible_devices", "string[]", "Y", "가입 가능 기기", '["Galaxy S series", "Galaxy Tab series"]', "Basic은 폴더블 미포함"],
    ["excluded_devices", "string[]", "Y", "가입 불가 기기", '["Galaxy Z Fold series"]', "Basic만 해당"],
    ["response_time.mail_in_days", "number", "N", "택배 수리 소요일", "5 / 3", ""],
    ["response_time.walk_in_hours", "number", "N", "방문 수리 소요 시간", "2 / 1", ""],
    ["response_time.remote_support_minutes", "number", "N", "원격 지원 소요 분", "15 / 5", ""],
    ["response_time.replacement_hours", "number", "N", "즉시 교체 소요 시간 (Premium만)", "- / 24", ""],
]
for row in care_data:
    ws6.append(row)
style_cells(ws6, 2, len(care_data) + 1, len(headers6))

# Device pricing
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="기기 카테고리별 월 보험료 (pricing_by_device_category)")
ws6.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["카테고리", "대표 기기", "Basic", "Premium", "비고", ""], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 6)

dev_price = [
    ["smartphone_flagship", "S25 Ultra, S25+, S24 Ultra, S24+", "$11.99", "$17.99", ""],
    ["smartphone_standard", "S25, S24, A55, A35, A25", "$8.99", "$13.99", ""],
    ["smartphone_foldable", "Z Fold6, Z Fold5, Z Flip6, Z Flip5", "N/A", "$17.99", "Premium 전용"],
    ["tablet", "Tab S10 Ultra~S9", "$7.99", "$12.99", ""],
    ["watch", "Watch Ultra, Watch7, Watch6", "$3.99", "$5.99", ""],
    ["earbuds", "Buds3 Pro, Buds3, Buds2 Pro, Buds FE", "$2.99", "$4.99", ""],
    ["notebook", "Book4 Ultra, Book4 Pro, Book4", "$9.99", "$14.99", ""],
]
start_r = r + 1
for row in dev_price:
    r += 1
    for i, v in enumerate(row, 1):
        ws6.cell(row=r, column=i, value=v)
style_cells(ws6, start_r, r, 6)

# Late enrollment
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="Late Enrollment 조건 (enrollment_rules.late_enrollment)")
ws6.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["항목", "값", "설명", "", "", ""], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 6)

late_data = [
    ["standard_window_days", "60", "일반 가입 기간 (구매 후 60일 이내)"],
    ["max_days_after_purchase", "365", "Late Enrollment 최대 허용일"],
    ["late_enrollment_fee", "$49", "수수료 (프로모션 시 면제 가능)"],
    ["required_condition", "good", "최소 요구 상태"],
    ["condition_requirements.screen", '["no_scratches", "light_scratches"]', "화면: 양호 이상만 통과"],
    ["condition_requirements.body", '["pristine", "minor_wear"]', "외관: 양호 이상만 통과"],
    ["condition_requirements.camera", '["clear", "minor_smudge"]', "카메라: 양호 이상만 통과"],
    ["vision_check_process", "4단계", "사진 업로드 → AI 분석 → 통과/불합격 → 가입/센터 방문"],
]
start_r = r + 1
for row in late_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws6.cell(row=r, column=i, value=v)
style_cells(ws6, start_r, r, 6)

# Claim process
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="클레임 프로세스 (claim_process) — 4단계")
ws6.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["단계", "타이틀", "설명", "소요 시간", "채널", ""], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 6)

claim_data = [
    [1, "클레임 접수", "Samsung Members 앱/웹/전화로 접수", "-", "앱, samsung.com, 1-800-SAMSUNG"],
    [2, "접수 확인", "확인 이메일 발송, 추가 정보 필요 시 연락", "30분", "-"],
    [3, "수리/교체 진행", "택배 수거(무료) 또는 서비스 센터 방문", "-", "mail_in, walk_in, instant_replacement"],
    [4, "완료 및 수령", "수리 완료 후 무료 배송, 교체 시 기존 기기 반납", "3~5일", "-"],
]
start_r = r + 1
for row in claim_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws6.cell(row=r, column=i, value=v)
style_cells(ws6, start_r, r, 6)

# Service centers
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="서비스 센터 (service_centers)")
ws6.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["지역", "센터 수", "워크인 가능", "", "", ""], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 6)

sc_data = [
    ["US", 420, "Y"],
    ["KR", 180, "Y"],
    ["UK", 85, "Y"],
    ["DE", 120, "Y"],
    ["JP", 95, "Y"],
    ["SG", 25, "Y"],
    ["전체", 3200, "-"],
]
start_r = r + 1
for row in sc_data:
    r += 1
    for i, v in enumerate(row, 1):
        ws6.cell(row=r, column=i, value=v)
style_cells(ws6, start_r, r, 6)

# Exclusions
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="보장 제외 사항 (exclusions)")
ws6.cell(row=r, column=1).font = subheader_font
exclusions = [
    "고의적 파손 또는 악의적 사용",
    "비공인 서비스 센터에서의 수리 이력",
    "루팅, 탈옥 등 소프트웨어 개조",
    "자연재해 (홍수, 지진, 화재 등)",
    "전쟁, 테러 등 불가항력",
    "미용 목적의 수리 (기능에 영향 없는 스크래치 등)",
    "분실/도난 보장은 Premium 플랜만 해당",
]
for exc in exclusions:
    r += 1
    ws6.cell(row=r, column=1, value="•  " + exc).font = cell_font

# FAQ
ws6.append([])
r = ws6.max_row + 1
ws6.cell(row=r, column=1, value="FAQ (10건, 전체 한글)")
ws6.cell(row=r, column=1).font = subheader_font
r += 1
for i, h in enumerate(["ID", "카테고리", "질문", "답변 (요약)", "", ""], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 6)

care_faq = [
    ["care-faq-001", "enrollment", "가입 후 언제부터 보장?", "즉시 (30일 내 추가 검증 가능)"],
    ["care-faq-002", "coverage", "해외 서비스 가능?", "Premium만, Samsung 센터 소재국"],
    ["care-faq-003", "billing", "가입 취소 방법?", "언제든 해지, 잔여 기간 환불"],
    ["care-faq-004", "plans", "Basic vs Premium 차이?", "분실/도난, 즉시교체, 화면$0, 힌지, 해외"],
    ["care-faq-005", "enrollment", "Late Enrollment이란?", "60일~1년, Vision 검사, 수수료 $49"],
    ["care-faq-006", "claims", "클레임 횟수 제한?", "연 2회, 분실/도난 별도 2회(Premium)"],
    ["care-faq-007", "coverage", "보장 안 되는 수리?", "고의파손, 루팅, 비공인수리, 자연재해"],
    ["care-faq-008", "claims", "교체 기기는 새 제품?", "공인 리퍼비시 또는 새 제품"],
    ["care-faq-009", "coverage", "Galaxy Club 가입 시 Care+ 필요?", "Galaxy Club에 Care+ 포함"],
    ["care-faq-010", "foldable", "폴더블 힌지 보장?", "Premium만, 자기부담금 $0"],
]
start_r = r + 1
for row in care_faq:
    r += 1
    for i, v in enumerate(row, 1):
        ws6.cell(row=r, column=i, value=v)
style_cells(ws6, start_r, r, 6)

auto_width(ws6, len(headers6))

# ============================================================
# Sheet 7: 확인 요청 사항
# ============================================================
ws7 = wb.create_sheet("확인 요청 사항")

headers7 = ["#", "항목", "질문", "현재 Mock 기준", "비고"]
ws7.append(headers7)
style_header(ws7, 1, len(headers7))

confirm_data = [
    [1, "데이터 제공 형식", "JSON / CSV / API 중 어떤 형태로 제공 가능한지?", "JSON 파일 4개 (정적)", "API 연동 시 엔드포인트 스펙 필요"],
    [2, "대상 지역", "실제 서비스 대상 지역은? (가격/통화/세금 영향)", "US, KR, UK, DE, JP, SG (6개국)", "현재 6개국 통화 변환 구현 완료"],
    [3, "대상 기기 범위", "Trade-in / Galaxy Club / Care+ 포함 기기 라인업 확정", "42개 기기 (S22~S25, Z4~6, Tab, Watch, Buds, Book, A, Ring)", "기기 추가/제거 시 전체 데이터 영향"],
    [4, "Trade-in 가격 정책", "기준가(base_values)/계수/감가 테이블이 실제 정책과 일치하는지?", "DATA_SPEC.md 기준 Mock 값", "가격 산출 공식 검증 필요"],
    [5, "프로모션 업데이트 주기", "정적 데이터 vs 실시간 API 연동?", "정적 JSON (22개 캠페인)", "실시간 시 API 스펙 필요"],
    [6, "Care+ Late Enrollment", "Vision 기반 상태 검사가 실제 서비스에 포함되는지?", "ChatGPT Vision → MCP 도구 호출 흐름 구현 완료", "데모 핵심 기능"],
    [7, "Galaxy Club 비용 비교", "기기별 구독 vs 일반구매 비교 데이터 산출 기준 확인", "7건 비용 비교 데이터 (Trade-in 시나리오 포함)", "위젯에서 12~36개월 슬라이더로 시뮬레이션"],
    [8, "통화/환율", "환율은 고정값 사용 vs 실시간 조회?", "고정 환율 (region_multipliers 내 base_rate)", "서버 fmtPrice() + 위젯 fmtP()로 변환"],
    [9, "이미지 리소스", "기기/프로모션 이미지 URL 또는 파일 제공 여부", "placeholder URL 사용 중", "위젯 렌더링에 필요"],
    [10, "데이터 갱신 주기", "각 데이터의 갱신 주기 (기기: 분기? 프로모션: 수시?)", "정적 (갱신 없음)", "운영 계획에 따라 결정"],
    [11, "서비스센터 데이터", "Care+ 서비스센터 위치/수 데이터 실제 값 확인", "Mock: 전체 3,200개 (US 420, KR 180 등)", "클레임 프로세스에서 참조"],
    [12, "Galaxy Club 가입자 수", "플랜별 가입자 수 데이터가 실제 수치인지?", "Mock: Basic 48,230 / Premium 31,450 / Family 12,780", "위젯에서 미표시, 내부 참고용"],
]
for row in confirm_data:
    ws7.append(row)
style_cells(ws7, 2, len(confirm_data) + 1, len(headers7))

auto_width(ws7, len(headers7))

# Save
output_path = r"C:\Users\user\samsung-galaxy-mcp-share\MCP_데이터_요청_목록.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
